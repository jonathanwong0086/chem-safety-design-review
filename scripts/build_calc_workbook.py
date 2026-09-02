#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""按输入 JSON 生成“活公式”计算复核 Excel 工作簿（.xlsx）。

用法：
    python scripts/build_calc_workbook.py 复核输入.json 计算复核.xlsx
    python scripts/build_calc_workbook.py --demo 计算复核.xlsx   # 用内置示例数据

产出的工作簿是“活的”：黄色单元格是可改输入（q、Q、体积等），
公式列写的是真正的 Excel 公式（如 =C3/D3），用 Excel 或 LibreOffice 打开会
自动重算并刷新“判定”列。这样审查方可以当场改数验证，说服力强于一段文字结论。

本脚本只依赖 openpyxl（本机已装 3.1.5），不需要 pandas / Excel。

输入 JSON 结构（各段都可选，缺了就不出对应工作表）：
{
  "重大危险源": [
    {"单元": "2#罐组", "物料": "溴", "q": 62, "Q": 20, "文档qQ": 1.55},
    ...
  ],
  "泄压面积": [
    {"区域": "105车间一层", "V": 3507.5, "C": 0.11, "文档A": 39.09},
    ...
  ],
  "物料平衡": {
    "工序": "三乙胺回收",
    "进料": [{"物料": "含三乙胺盐酸盐水相", "量": 2786}, ...],
    "出料": [{"物料": "回收三乙胺", "量": 590}, ...]
  }
}

约定：黄色填充＝可改输入格；浅灰填充＝公式格（自动算）；表头标 [输入] / [=公式]。
"""
import json
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.comments import Comment
except ImportError:
    sys.stderr.write("缺少 openpyxl，请先安装：pip install openpyxl\n")
    sys.exit(1)

# 样式常量
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")   # 黄色：可改输入
FORMULA_FILL = PatternFill("solid", fgColor="E8E8E8")  # 浅灰：公式格
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # 深蓝表头
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")      # 红：偏差判定不符
HEADER_FONT = Font(name="宋体", bold=True, color="FFFFFF")
BODY_FONT = Font(name="宋体")
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def _cell(ws, row, col, value=None, fill=None, formula=None, note=None):
    """写一个单元格：value 是静态值，formula 是公式串（以 = 开头）。"""
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
        c.fill = fill or FORMULA_FILL
    else:
        c.value = value
        if fill is not None:
            c.fill = fill
    c.font = BODY_FONT
    c.border = BORDER
    c.alignment = CENTER
    if note:
        c.comment = Comment(note, "chem-safety")
    return c
def build_msds_sheet(wb, rows):
    """重大危险源 q/Q 复核表：辨识层 S=Σ(q/Q)。黄格为 q、Q 输入，E 列自动算 q/Q。"""
    ws = wb.create_sheet("重大危险源qQ")
    headers = ["单元", "物料", "q 最大存在量(t) [输入]", "Q 临界量(t) [输入]",
               "q/Q [=公式]", "文档q/Q [输入]", "偏差判定 [=公式]"]
    for j, h in enumerate(headers, start=1):
        _cell(ws, 1, j, h)
        _style_header(ws.cell(row=1, column=j))
    r = 2
    for item in rows:
        _cell(ws, r, 1, item.get("单元", ""))
        _cell(ws, r, 2, item.get("物料", ""))
        _cell(ws, r, 3, item.get("q"), fill=INPUT_FILL)
        _cell(ws, r, 4, item.get("Q"), fill=INPUT_FILL)
        _cell(ws, r, 5, formula='=IF(D%d=0,"",C%d/D%d)' % (r, r, r))
        _cell(ws, r, 6, item.get("文档qQ"), fill=INPUT_FILL)
        _cell(ws, r, 7, formula='=IF(F%d="","—",IF(ABS(E%d-F%d)<=0.01,"一致","★偏差"))' % (r, r, r))
        r += 1
    _cell(ws, r, 1, "Σq/Q（全表合计）")
    _cell(ws, r, 5, formula="=SUM(E2:E%d)" % (r - 1))
    _cell(ws, r, 7, formula='=IF(E%d>=1,"构成重大危险源(S>=1)","不构成(S<1)")' % r)
    for col, w in zip("ABCDEFG", (14, 22, 20, 18, 14, 14, 24)):
        ws.column_dimensions[col].width = w
    return ("重大危险源 q/Q（辨识层 S=Σq/Q）", "重大危险源qQ", "E%d" % r)


def build_vent_sheet(wb, rows):
    """泄压面积复核：A = 10 * C * V^(2/3)。黄格为 V、C，D 列自动算 A。"""
    ws = wb.create_sheet("泄压面积")
    headers = ["区域", "体积V(m³) [输入]", "泄压比C [输入]",
               "A计算(m²) [=公式]", "A文档(m²) [输入]", "判定 [=公式]"]
    for j, h in enumerate(headers, start=1):
        _cell(ws, 1, j, h)
        _style_header(ws.cell(row=1, column=j))
    r = 2
    for item in rows:
        _cell(ws, r, 1, item.get("区域", ""))
        _cell(ws, r, 2, item.get("V"), fill=INPUT_FILL)
        _cell(ws, r, 3, item.get("C", 0.11), fill=INPUT_FILL)
        _cell(ws, r, 4, formula="=10*C%d*B%d^(2/3)" % (r, r))
        _cell(ws, r, 5, item.get("文档A"), fill=INPUT_FILL)
        _cell(ws, r, 6, formula='=IF(E%d="","—",IF(ABS(D%d-E%d)/D%d<=0.05,"一致","★不符"))' % (r, r, r, r))
        r += 1
    for col, w in zip("ABCDEF", (26, 16, 14, 16, 16, 16)):
        ws.column_dimensions[col].width = w
    return ("泄压面积 A=10·C·V^(2/3)", "泄压面积", "D2")


def build_balance_sheet(wb, spec):
    """物料平衡闭合：闭合差 = 进料合计 - 出料合计。黄格为各物料量。"""
    ws = wb.create_sheet("物料平衡闭合")
    title = "%s 物料平衡闭合复核" % spec.get("工序", "")
    _cell(ws, 1, 1, title)
    _style_header(ws.cell(row=1, column=1))
    for j in range(2, 5):
        _style_header(ws.cell(row=1, column=j))
    _cell(ws, 2, 1, "进料物料")
    _cell(ws, 2, 2, "量(t/批) [输入]")
    _cell(ws, 2, 3, "出料物料")
    _cell(ws, 2, 4, "量(t/批) [输入]")
    for j in range(1, 5):
        _style_header(ws.cell(row=2, column=j))
    feeds = spec.get("进料", [])
    outs = spec.get("出料", [])
    n = max(len(feeds), len(outs))
    r0 = 3
    for i in range(n):
        r = r0 + i
        if i < len(feeds):
            _cell(ws, r, 1, feeds[i].get("物料", ""))
            _cell(ws, r, 2, feeds[i].get("量"), fill=INPUT_FILL)
        else:
            _cell(ws, r, 1, "")
            _cell(ws, r, 2, None)
        if i < len(outs):
            _cell(ws, r, 3, outs[i].get("物料", ""))
            _cell(ws, r, 4, outs[i].get("量"), fill=INPUT_FILL)
        else:
            _cell(ws, r, 3, "")
            _cell(ws, r, 4, None)
    rt = r0 + n
    _cell(ws, rt, 1, "进料合计")
    _cell(ws, rt, 2, formula="=SUM(B%d:B%d)" % (r0, rt - 1))
    _cell(ws, rt, 3, "出料合计")
    _cell(ws, rt, 4, formula="=SUM(D%d:D%d)" % (r0, rt - 1))
    rd = rt + 1
    _cell(ws, rd, 1, "闭合差 = 进料 - 出料")
    _cell(ws, rd, 2, formula="=B%d-D%d" % (rt, rt))
    _cell(ws, rd, 3, "判定")
    _cell(ws, rd, 4, formula='=IF(ABS(B%d)<=0.01*B%d,"闭合","★不平")' % (rd, rt))
    for col, w in zip("ABCD", (24, 16, 24, 16)):
        ws.column_dimensions[col].width = w
    return ("物料平衡闭合（进料-出料）", "物料平衡闭合", "B%d" % rd)


def build_summary_sheet(wb, refs):
    """自校核汇总：把各表的关键判定单元格用公式引过来，一屏总览。"""
    ws = wb.create_sheet("自校核汇总", 0)
    _cell(ws, 1, 1, "关键计算自校核汇总（公式联动，改输入自动刷新）")
    _cell(ws, 1, 2, "")
    _cell(ws, 1, 3, "")
    for j in range(1, 4):
        _style_header(ws.cell(row=1, column=j))
    _cell(ws, 2, 1, "检查项")
    _cell(ws, 2, 2, "关联单元格")
    _cell(ws, 2, 3, "结果 [=公式]")
    for j in range(1, 4):
        _style_header(ws.cell(row=2, column=j))
    r = 3
    for label, sheet, cellref in refs:
        _cell(ws, r, 1, label)
        _cell(ws, r, 2, "%s!%s" % (sheet, cellref))
        _cell(ws, r, 3, formula="='%s'!%s" % (sheet, cellref))
        r += 1
    _cell(ws, r, 1, "说明：黄色格为可编辑输入；改后 Excel/LibreOffice 自动重算各判定列。")
    for col, w in zip("ABC", (36, 24, 28)):
        ws.column_dimensions[col].width = w
    return ws


DEMO = {
    "重大危险源": [
        {"单元": "2#罐组", "物料": "溴(2台一用一应急)", "q": 62, "Q": 20, "文档qQ": 1.55},
        {"单元": "2#罐组", "物料": "三氯化磷", "q": 47.1, "Q": 500, "文档qQ": 0.0942},
        {"单元": "105车间", "物料": "三乙胺精馏釜", "q": 2.5, "Q": 1000, "文档qQ": 0.25},
    ],
    "泄压面积": [
        {"区域": "105车间一层上料间", "V": 3507.5, "C": 0.11, "文档A": 39.09},
        {"区域": "1#仓库区3", "V": 1721.37, "C": 0.11, "文档A": 158.0},
    ],
    "物料平衡": {
        "工序": "三乙胺回收",
        "进料": [{"物料": "含三乙胺盐酸盐水相", "量": 2786}, {"物料": "30%液碱", "量": 1086}],
        "出料": [{"物料": "回收三乙胺", "量": 590}, {"物料": "废水", "量": 3910}],
    },
}


def build_workbook(data, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删掉默认表
    refs = []
    if data.get("重大危险源"):
        refs.append(build_msds_sheet(wb, data["重大危险源"]))
    if data.get("泄压面积"):
        refs.append(build_vent_sheet(wb, data["泄压面积"]))
    if data.get("物料平衡"):
        refs.append(build_balance_sheet(wb, data["物料平衡"]))
    if not refs:
        sys.stderr.write("输入里没有任何可复核的段（重大危险源/泄压面积/物料平衡），未生成。\n")
        sys.exit(2)
    build_summary_sheet(wb, refs)
    wb.save(out_path)
    return len(refs)


def main(argv):
    if len(argv) < 2:
        sys.stderr.write(__doc__)
        sys.exit(1)
    if argv[0] == "--demo":
        data, out_path = DEMO, argv[1]
    else:
        if len(argv) < 2:
            sys.stderr.write("用法：build_calc_workbook.py 输入.json 输出.xlsx（或 --demo 输出.xlsx）\n")
            sys.exit(1)
        with open(argv[0], encoding="utf-8") as f:
            data = json.load(f)
        out_path = argv[1]
    n = build_workbook(data, out_path)
    sys.stdout.write("已生成计算复核工作簿：%s（含 %d 个复核表 + 自校核汇总）\n" % (out_path, n))


if __name__ == "__main__":
    main(sys.argv[1:])
